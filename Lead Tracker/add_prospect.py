"""
Called by Claude after analyzing a prospect screenshot.
Adds one row to the tracker and generates DMs immediately.
"""
import datetime, sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

TRACKER = "C:/Users/diazc/OneDrive/Desktop/FJDMedia/FJDMedia Website/Lead Tracker/FJMedia_Lead_Tracker.xlsx"
DM_OUT  = "C:/Users/diazc/OneDrive/Desktop/FJDMedia/FJDMedia Website/Lead Tracker/DMs_Latest.html"

NAVY       = "071520"
GOLD       = "C9A84C"
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F7FA"
MID_GREY   = "DCE1E8"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────────────────────────────────────
# DM generation logic — James's tone
# ─────────────────────────────────────────────────────────────────────────────
def generate_dms(name, business, niche, contact, source, has_website, followers, hook):
    biz    = business or contact
    handle = contact if contact.startswith("@") else f"@{contact}"

    # Build the opener based on source
    openers = {
        "Instagram":    f"Came across {biz} on IG",
        "WPG Explore":  f"Found you guys through WPG Explore",
        "Comment":      f"Saw the comment on my post — appreciate it",
        "Referral":     f"Heard about {biz} through a mutual",
        "Word of Mouth":f"Heard about {biz} through a mutual",
    }
    opener = openers.get(source, f"Came across {biz} on IG")

    # Presence line
    if has_website == "No":
        pres_q = "do you have a website or is everything just running through IG right now?"
        pres_context = "no website anywhere"
    elif has_website == "Unknown":
        pres_q = "do you have a website somewhere or is it all social media?"
        pres_context = "unclear web presence"
    else:
        pres_q = "when was the last time your website got updated?"
        pres_context = "has a website"

    # Hook line (what stood out)
    hook_line = f" {hook.rstrip('.')}." if hook else "."

    dm1 = f"Hey, {opener} --{hook_line}\n\nQuick question, {pres_q}\n\n- James"

    dm2 = f"Hey -- {opener.lower()}.{' ' + hook.rstrip('.') + '.' if hook else ''}\n\nRandom but {pres_q}\n\n- James"

    dm3 = (
        f"Hey, {opener} --{hook_line}\n\n"
        f"Wanted to reach out -- {pres_q} "
        f"I build websites for local Winnipeg businesses. "
        f"I put together a full preview before anyone pays anything.\n\n- James"
    )

    return [
        ("Option 1 — Casual curiosity", dm1),
        ("Option 2 — Short & punchy",   dm2),
        ("Option 3 — More context",     dm3),
    ]

# ─────────────────────────────────────────────────────────────────────────────
# HTML output
# ─────────────────────────────────────────────────────────────────────────────
def write_html(business, contact, dms):
    options_html = ""
    for label, msg in dms:
        escaped = msg.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace("\n","<br>")
        options_html += f"""
        <div class="dm-card">
          <div class="label">{label}</div>
          <div class="message">{escaped}</div>
          <button onclick="copy(this, `{msg.replace('`','\\`')}`)">Copy</button>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>DMs — {business}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ background:#f5f7fa; font-family:'Segoe UI',sans-serif; color:#071520; padding:40px 20px; }}
  .wrap {{ max-width:680px; margin:0 auto; }}
  header {{ margin-bottom:28px; }}
  .logo {{ font-family:'Courier New',monospace; font-size:20px; font-weight:bold; }}
  .logo span {{ color:#c9a84c; }}
  h2 {{ font-size:14px; color:#444; margin-top:4px; }}
  .dm-card {{ background:#fff; border:1px solid #dce1e8; border-radius:12px; padding:22px; margin-bottom:16px; }}
  .label {{ font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:.06em; color:#c9a84c; margin-bottom:10px; }}
  .message {{ font-size:14px; line-height:1.75; color:#071520; }}
  button {{ margin-top:14px; padding:8px 18px; background:#071520; color:#c9a84c; border:none;
            border-radius:7px; font-size:12px; font-weight:700; cursor:pointer; letter-spacing:.04em; }}
  button:hover {{ background:#0d1f3c; }}
  button.copied {{ background:#2e7d52; color:#fff; }}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <div class="logo"><span>{{FJ}}</span> Media — DM Generator</div>
    <h2>Prospect: {business} ({contact})</h2>
  </header>
  {options_html}
</div>
<script>
function copy(btn, text) {{
  navigator.clipboard.writeText(text).then(() => {{
    btn.textContent = 'Copied';
    btn.classList.add('copied');
    setTimeout(() => {{ btn.textContent = 'Copy'; btn.classList.remove('copied'); }}, 2000);
  }});
}}
</script>
</body>
</html>"""
    with open(DM_OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"DMs written: {DM_OUT}")

# ─────────────────────────────────────────────────────────────────────────────
# Main — call this with prospect data
# ─────────────────────────────────────────────────────────────────────────────
def add_prospect(
    name, business, niche, contact, source,
    has_website="No", followers="", hook="",
    package="TBD", follow_up="", notes=""
):
    wb  = load_workbook(TRACKER)
    ws  = wb["Leads"]
    today     = datetime.date.today().strftime("%Y-%m-%d")
    next_row  = ws.max_row + 1
    bg        = WHITE if next_row % 2 == 0 else LIGHT_GREY

    values = [
        name, business, niche, contact, source,
        "Cold", today, follow_up, package,
        has_website, followers, hook,
        "", "", notes
    ]

    for c_idx, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=c_idx, value=val)
        cell.fill      = fill(bg)
        cell.font      = Font(name="Segoe UI", size=10, color=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=(c_idx in (12, 15)))
        cell.border    = thin_border()

    ws.row_dimensions[next_row].height = 22
    wb.save(TRACKER)
    print(f"Logged: {business} ({contact}) — row {next_row}")

    dms = generate_dms(name, business, niche, contact, source, has_website, followers, hook)
    write_html(business, contact, dms)

    print("\n-- DM Options --------------------------------------")
    for label, msg in dms:
        print(f"\n[{label}]\n{msg}\n")

# ─────────────────────────────────────────────────────────────────────────────
# To add a new prospect, Claude edits the call below and runs this file
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    add_prospect(
        name        = "Anna",
        business    = "Ms. Groomer",
        niche       = "Pet Grooming",
        contact     = "@ms_groomer",
        source      = "WPG Explore",
        has_website = "No",
        followers   = "86",
        hook        = "Breed-standard grooming, 232 posts, solid before/afters — doing real work but invisible online",
        package     = "TBD",
        follow_up   = "",
        notes       = "No website, no booking page, no directory listing. IG only confirmed via web search.",
    )
