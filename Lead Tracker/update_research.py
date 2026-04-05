"""
Updates existing prospects with research findings and generates DMs for all.
"""
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

TRACKER = "C:/Users/diazc/OneDrive/Desktop/FJDMedia/FJDMedia Website/Lead Tracker/FJMedia_Lead_Tracker.xlsx"
DM_OUT  = "C:/Users/diazc/OneDrive/Desktop/FJDMedia/FJDMedia Website/Lead Tracker/DMs_Latest.html"

NAVY       = "071520"
GOLD       = "C9A84C"
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F7FA"
MID_GREY   = "DCE1E8"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

# Research results for each prospect
UPDATES = {
    "Gold Prime Renovations": {
        "has_website": "No",
        "followers": "341",
        "hook": "Solid portfolio of bathrooms, basements and commercial builds. Actively hiring so they're growing.",
        "notes": "Facebook only (238 likes). No website, no Google Business Profile, no directory listings. Growing company — hiring means they need to look more credible.",
    },
    "MB Contracting": {
        "has_website": "No",
        "followers": "",
        "hook": "No web presence found anywhere — completely invisible online.",
        "notes": "No website, no Google Business Profile, no directory listings. IG only. Not even on BBB or Yellow Pages.",
    },
    "Elara Petals": {
        "has_website": "No",
        "followers": "257",
        "hook": "Unique ribbon bouquet niche, consistent content, custom orders — but only reachable through DMs.",
        "notes": "No website, no Etsy, no Google Business Profile. IG only confirmed via web search. Missing out on anyone who searches for custom bouquets in Winnipeg.",
    },
}

def generate_dms(business, contact, source, has_website, hook):
    openers = {
        "Instagram":     f"Came across {business} on IG",
        "WPG Explore":   f"Found you guys through WPG Explore",
        "Comment":       f"Saw the comment on my post",
        "Referral":      f"Heard about {business} through a mutual",
        "Word of Mouth": f"Heard about {business} through a mutual",
    }
    opener = openers.get(source, f"Came across {business} on IG")

    if has_website == "No":
        pres_q = "do you have a website or is everything just running through IG right now?"
    else:
        pres_q = "do you have a website somewhere or is it all social media?"

    hook_line = f" {hook.rstrip('.')}." if hook else "."

    dm1 = f"Hey, {opener} --{hook_line}\n\nQuick question, {pres_q}\n\n- James"
    dm2 = f"Hey -- {opener.lower()}.{' ' + hook.rstrip('.') + '.' if hook else ''}\n\nRandom but {pres_q}\n\n- James"
    dm3 = (
        f"Hey, {opener} --{hook_line}\n\n"
        f"Wanted to reach out -- {pres_q} "
        f"I build websites for local Winnipeg businesses. "
        f"I put together a full preview before anyone pays anything.\n\n- James"
    )
    return [
        ("Option 1 -- Casual curiosity", dm1),
        ("Option 2 -- Short & punchy",   dm2),
        ("Option 3 -- More context",     dm3),
    ]

def update_tracker():
    wb = load_workbook(TRACKER)
    ws = wb["Leads"]

    # Column index map (1-based)
    # J=10 has_website, K=11 followers, L=12 hook, O=15 notes
    COL = {"has_website": 10, "followers": 11, "hook": 12, "notes": 15}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        business = row[1].value  # Column B
        if business in UPDATES:
            data = UPDATES[business]
            for field, col_idx in COL.items():
                cell = ws.cell(row=row[0].row, column=col_idx)
                cell.value = data[field]

    wb.save(TRACKER)
    print(f"Tracker updated.")

def write_html(all_prospects):
    sections = ""
    for biz, contact, source, has_website, hook, dms in all_prospects:
        options_html = ""
        for label, msg in dms:
            escaped = msg.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace("\n","<br>")
            js_msg  = msg.replace("\\","\\\\").replace("`","\\`").replace("$","\\$")
            options_html += (
                '<div class="dm-card">'
                f'<div class="label">{label}</div>'
                f'<div class="message">{escaped}</div>'
                f'<button onclick="copy(this, `{js_msg}`)">Copy</button>'
                '</div>'
            )

        sections += f"""
        <div class="prospect-block">
          <div class="prospect-header">{biz} <span>{contact}</span></div>
          {options_html}
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>DMs -- FJMedia</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ background:#f5f7fa; font-family:'Segoe UI',sans-serif; color:#071520; padding:40px 20px; }}
  .wrap {{ max-width:720px; margin:0 auto; }}
  header {{ margin-bottom:32px; }}
  .logo {{ font-family:'Courier New',monospace; font-size:20px; font-weight:bold; }}
  .logo span {{ color:#c9a84c; }}
  header p {{ font-size:13px; color:#666; margin-top:4px; }}
  .prospect-block {{ margin-bottom:40px; }}
  .prospect-header {{ font-size:16px; font-weight:700; color:#071520; margin-bottom:14px;
                      padding-bottom:10px; border-bottom:2px solid #c9a84c; }}
  .prospect-header span {{ font-size:13px; font-weight:400; color:#888; margin-left:8px; }}
  .dm-card {{ background:#fff; border:1px solid #dce1e8; border-radius:12px; padding:22px; margin-bottom:12px; }}
  .label {{ font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:.06em; color:#c9a84c; margin-bottom:10px; }}
  .message {{ font-size:14px; line-height:1.75; color:#071520; }}
  button {{ margin-top:14px; padding:8px 18px; background:#071520; color:#c9a84c; border:none;
            border-radius:7px; font-size:12px; font-weight:700; cursor:pointer; letter-spacing:.04em; }}
  button:hover {{ background:#0d1f3c; }}
  button.copied {{ background:#2e7d52; color:#fff; }}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <div class="logo"><span>{{FJ}}</span> Media -- DM Generator</div>
    <p>All prospects -- pick your message, copy, send.</p>
  </header>
  {sections}
</div>
<script>
function copy(btn, text) {{
  navigator.clipboard.writeText(text).then(() => {{
    btn.textContent = 'Copied';
    btn.classList.add('copied');
    setTimeout(() => {{ btn.textContent = 'Copy'; btn.classList.remove('copied'); }}, 2000);
  }});
}}
</script>
</body>
</html>"""

    with open(DM_OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"DMs written: {DM_OUT}")

if __name__ == "__main__":
    update_tracker()

    # Generate DMs for all 3 using updated data
    prospects_for_dm = [
        ("Gold Prime Renovations", "@goldprimerenovations", "Instagram", "No",
         "Solid portfolio of bathrooms, basements and commercial builds. Actively hiring so they're growing."),
        ("MB Contracting", "@mbgeneralcontracting", "WPG Explore", "No",
         "No web presence at all -- completely invisible online despite being an active contracting company."),
        ("Elara Petals", "@elarapetals", "WPG Explore", "No",
         "Unique ribbon bouquet niche, consistent content, custom orders -- but only reachable through DMs."),
    ]

    all_output = []
    for biz, contact, source, has_web, hook in prospects_for_dm:
        dms = generate_dms(biz, contact, source, has_web, hook)
        all_output.append((biz, contact, source, has_web, hook, dms))

    write_html(all_output)
    print("Done.")
