"""
Rebuilds the tracker with expanded columns for DM generation.
Migrates all existing data. Run once.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

OUTPUT = "C:/Users/diazc/OneDrive/Desktop/FJDMedia/FJDMedia Website/Lead Tracker/FJMedia_Lead_Tracker.xlsx"

NAVY       = "071520"
GOLD       = "C9A84C"
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F7FA"
MID_GREY   = "DCE1E8"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

HEADERS = [
    "Name",            # A
    "Business",        # B
    "Niche",           # C
    "Contact",         # D  IG handle / email
    "Source",          # E
    "Stage",           # F
    "Last Touchpoint", # G
    "Follow-Up Date",  # H
    "Package Interest",# I
    "Has Website",     # J  Yes / No / Unknown
    "Followers",       # K
    "Hook",            # L  what stood out — used for DM personalization
    "DM Sent",         # M  date sent
    "DM Response",     # N  what they said back
    "Notes",           # O
]

COL_WIDTHS = {
    "A": 18, "B": 26, "C": 18, "D": 24, "E": 16,
    "F": 16, "G": 20, "H": 18, "I": 24, "J": 14,
    "K": 12, "L": 36, "M": 14, "N": 30, "O": 40,
}

today = datetime.date.today().strftime("%Y-%m-%d")

SEED_DATA = [
    # Name, Business, Niche, Contact, Source, Stage, Last Touchpoint, Follow-Up,
    # Package, Has Website, Followers, Hook, DM Sent, DM Response, Notes
    [
        "", "Gold Prime Renovations", "Renovations", "@goldprimerenovations",
        "Instagram", "Cold", today, "", "TBD",
        "No", "341", "Solid portfolio — bathrooms, basements, commercial builds. Actively hiring so they're growing.",
        "", "", "B2C pitch angle. Only Facebook in bio."
    ],
    [
        "", "MB Contracting", "Construction", "@mbgeneralcontracting",
        "WPG Explore", "Cold", today, "", "TBD",
        "No", "", "",
        "", "", "Found via WPG Explore feature."
    ],
    [
        "", "Elara Petals", "Gifts / Florals", "@elarapetals",
        "WPG Explore", "Cold", today, "", "TBD",
        "No", "257", "Ribbon bouquets, custom orders, consistent product content. Really unique niche.",
        "", "", "DM to order only. No booking page anywhere."
    ],
    [
        "Anna", "Ms. Groomer", "Pet Grooming", "@ms_groomer",
        "WPG Explore", "Cold", today, "", "TBD",
        "No", "86", "Breed-standard grooming, 232 posts, solid before/afters. Invisible online despite real skill.",
        "", "", "No website, no booking page, no directory listing. IG only confirmed via web search."
    ],
]

def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 32
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill      = fill(NAVY)
        cell.font      = Font(name="Segoe UI", bold=True, color=GOLD, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()

    # Dropdowns
    def dv(formula, col_range):
        d = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(d)
        d.sqref = col_range

    dv('"Cold,Contacted,Replied,Call Booked,Preview Sent,Approved,Paid,Dead,Revisit Later"', "F2:F1000")
    dv('"Instagram,WPG Explore,Referral,Cold DM,Comment,Word of Mouth,Other"', "E2:E1000")
    dv('"Get Online $300,Get Found $600,Get Customers $1000,Own the Market $1600,Retainer Maintain $150,Retainer Grow $300,TBD"', "I2:I1000")
    dv('"Yes,No,Unknown"', "J2:J1000")

    for r_idx, row in enumerate(SEED_DATA, 2):
        bg = WHITE if r_idx % 2 == 0 else LIGHT_GREY
        ws.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = fill(bg)
            cell.font      = Font(name="Segoe UI", size=10, color=NAVY)
            cell.alignment = Alignment(vertical="center", wrap_text=(c_idx in (12, 15)))
            cell.border    = thin_border()

    ws.freeze_panes = "A2"
    wb.save(OUTPUT)
    print(f"Rebuilt: {OUTPUT}")

if __name__ == "__main__":
    build()
