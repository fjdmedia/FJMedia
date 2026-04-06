from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

OUTPUT = "C:/Users/diazc/OneDrive/Desktop/FJDMedia/FJDMedia Website/Lead Tracker/FJMedia_Lead_Tracker.xlsx"

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY       = "071520"
GOLD       = "C9A84C"
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F7FA"
MID_GREY   = "DCE1E8"

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=MID_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {
        "A": 22,   # Name
        "B": 26,   # Business
        "C": 18,   # Niche
        "D": 24,   # Contact (IG handle / email)
        "E": 16,   # Source
        "F": 16,   # Stage
        "G": 20,   # Last Touchpoint
        "H": 18,   # Follow-Up Date
        "I": 24,   # Package Interest
        "J": 40,   # Notes
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ── Header row ────────────────────────────────────────────────────────────
    headers = [
        "Name", "Business", "Niche", "Contact",
        "Source", "Stage", "Last Touchpoint",
        "Follow-Up Date", "Package Interest", "Notes"
    ]
    ws.row_dimensions[1].height = 32

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill       = fill(NAVY)
        cell.font       = Font(name="Segoe UI", bold=True, color=GOLD, size=11)
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = thin_border()

    # ── Dropdown options ──────────────────────────────────────────────────────
    stage_dv = DataValidation(
        type="list",
        formula1='"Cold,Contacted,Replied,Call Booked,Preview Sent,Approved,Paid,Dead"',
        allow_blank=True, showDropDown=False
    )
    source_dv = DataValidation(
        type="list",
        formula1='"Instagram,WPG Explore,Referral,Cold DM,Comment,Word of Mouth,Other"',
        allow_blank=True, showDropDown=False
    )
    package_dv = DataValidation(
        type="list",
        formula1='"Get Online $300,Get Found $600,Get Customers $1000,Own the Market $1600,Retainer Maintain $150,Retainer Grow $300,TBD"',
        allow_blank=True, showDropDown=False
    )

    ws.add_data_validation(stage_dv)
    ws.add_data_validation(source_dv)
    ws.add_data_validation(package_dv)

    stage_dv.sqref   = "F2:F1000"
    source_dv.sqref  = "E2:E1000"
    package_dv.sqref = "I2:I1000"

    # ── Seed data — existing prospects ───────────────────────────────────────
    today = datetime.date.today().strftime("%Y-%m-%d")
    rows = [
        # Name, Business, Niche, Contact, Source, Stage, Last Touchpoint, Follow-Up, Package, Notes
        ["", "Gold Prime Renovations", "Renovations", "@goldprimerenovations", "Instagram", "Cold", today, "", "TBD", "No website. Solid portfolio — bathrooms, basements, commercial. Growing (hiring). B2C pitch angle."],
        ["", "MB Contracting",         "Construction", "@mbgeneralcontracting", "Instagram", "Cold", today, "", "TBD", "No website. Found via WPG Explore."],
        ["", "Elara Petals",           "Gifts / Florals", "@elarapetals",       "WPG Explore", "Cold", today, "", "TBD", "No website. Ribbon bouquets, custom orders via DM. Winnipeg-based. 257 followers, consistent content."],
        ["", "FAYNA Pekarnya",         "Bakery", "@fayna__pekarnya",             "Instagram", "Cold", today, "", "TBD", "No website (Canva link only). 321 followers. Commercial kitchen, licensed. Winnipeg — Nobside Cafe. Cakes, pastries, custom desserts. Already following."],
        ["", "Sugar + Salt Bakeshoppe","Bakery", "@sugarandsaltbakeshoppe",      "Instagram", "Cold", today, "", "TBD", "Has Squarespace site (sugarandsaltbakeshoppe.com) — redesign opportunity. 10.7K followers, 897 Corydon Ave. Sweet tables page has no booking form, no pricing, buried lead times. Pitch: better conversion + online ordering flow."],
        ["", "Gibson's Goodies",       "Bakery / Cookies", "@gibsons.goodies",   "Instagram", "Cold", today, "", "TBD", "No website. 362 followers. Self-taught cookie decorator. Email: gibsons.goodies.wpg@gmail.com. DM for orders. Winnipeg-based."],
        ["Rochelle", "For Sweet's Sake","Bakery / Cookies", "@forsweetssakewpg", "Instagram", "Cold", today, "", "TBD", "Linktree only (linktr.ee/forsweetssake). 2,644 followers, 409 posts. 'No DMs — email or form.' Strong content, zero web presence. Perfect full site candidate."],
    ]

    for r_idx, row in enumerate(rows, 2):
        bg = WHITE if r_idx % 2 == 0 else LIGHT_GREY
        ws.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = fill(bg)
            cell.font      = Font(name="Segoe UI", size=10, color=NAVY)
            cell.alignment = Alignment(vertical="center", wrap_text=(c_idx == 10))
            cell.border    = thin_border()

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")

if __name__ == "__main__":
    build()
